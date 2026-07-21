from __future__ import annotations

import csv
import hashlib
import json
import logging
import os
import re
import threading
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.config import BASE_DIR, get_settings
from app.services.ingestion import (
    inspect_te_file,
    iter_te_row_chunks,
    resolve_te_row_target,
    select_te_sample_indices,
    te_feature_names,
)

logger = logging.getLogger(__name__)

try:
    from docx import Document as DocxDocument
except Exception:  # noqa: BLE001
    DocxDocument = None

try:
    from pypdf import PdfReader
except Exception:  # noqa: BLE001
    PdfReader = None


SUPPORTED_EXTENSIONS = {
    ".csv",
    ".dat",
    ".docx",
    ".f",
    ".ini",
    ".json",
    ".md",
    ".pdf",
    ".txt",
    ".xlsx",
}
TEXT_EXTENSIONS = {".txt", ".md", ".ini", ".f"}
WHITESPACE_RE = re.compile(r"\s+")
OPENPYXL_WARNING_MODULE = r"openpyxl(\..*)?$"


@dataclass(slots=True)
class DocumentSection:
    document_name: str
    source: str
    section: str
    text: str
    page: str | None = None
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class DocumentChunk:
    chunk_id: str
    document_name: str
    source: str
    section: str
    page: str | None
    content: str
    metadata: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class _CachedFileMetadata:
    path: Path
    name: str
    suffix: str
    size: int = 0
    mtime_ns: int = 0
    resolved_source: str | None = None
    relative_source: str | None = None


@dataclass(slots=True)
class _DiscoverySnapshot:
    files: tuple[Path, ...]
    directories: tuple[Path, ...]
    directory_mtimes_ns: tuple[int, ...]


class TextChunker:
    def __init__(self, *, chunk_size: int, overlap: int):
        self.chunk_size = chunk_size
        self.overlap = overlap

    def chunk_sections(self, sections: list[DocumentSection]) -> list[DocumentChunk]:
        chunks: list[DocumentChunk] = []
        append_chunk = chunks.append
        normalize = self._normalize
        split = self._split
        for section in sections:
            normalized = normalize(section.text)
            if not normalized:
                continue
            metadata = section.metadata
            for index, content in enumerate(split(normalized)):
                chunk_id = hashlib.sha1(
                    f"{section.source}|{section.section}|{index}|{content}".encode("utf-8")
                ).hexdigest()
                append_chunk(
                    DocumentChunk(
                        chunk_id=chunk_id,
                        document_name=section.document_name,
                        source=section.source,
                        section=section.section,
                        page=section.page,
                        content=content,
                        metadata=dict(metadata),
                    )
                )
        return chunks

    def _split(self, text: str) -> list[str]:
        if len(text) <= self.chunk_size:
            return [text]
        chunks: list[str] = []
        append_chunk = chunks.append
        text_length = len(text)
        half_window = max(1, self.chunk_size // 2)
        start = 0
        while start < text_length:
            end = min(text_length, start + self.chunk_size)
            if end < text_length:
                split_at = text.rfind(" ", start + half_window, end)
                if split_at > start:
                    end = split_at
            chunk = text[start:end].strip()
            if chunk:
                append_chunk(chunk)
            if end >= text_length:
                break
            start = max(end - self.overlap, start + 1)
        return chunks

    @staticmethod
    def _normalize(value: str) -> str:
        return WHITESPACE_RE.sub(" ", value).strip()


class ModularDocumentLoader:
    def __init__(self, roots: list[Path] | None = None):
        settings = get_settings()
        self.settings = settings
        self.roots = roots or [BASE_DIR / "docs", settings.dataset_root]
        self._cache_lock = threading.Lock()
        self._discovery_cache: dict[str, _DiscoverySnapshot] = {}
        self._file_metadata_cache: dict[str, _CachedFileMetadata] = {}

    def discover_files(self) -> list[Path]:
        files: list[Path] = []
        for root in self.roots:
            files.extend(self._discover_root_files(Path(root)))
        files.sort()
        return files

    def load_sections(self, path: Path) -> list[DocumentSection]:
        suffix = path.suffix.lower()
        if suffix in TEXT_EXTENSIONS:
            return self._load_text(path)
        if suffix == ".json":
            return self._load_json(path)
        if suffix == ".csv":
            return self._load_csv(path)
        if suffix == ".xlsx":
            return self._load_xlsx(path)
        if suffix == ".dat":
            return self._load_te_dat(path)
        if suffix == ".pdf":
            return self._load_pdf(path)
        if suffix == ".docx":
            return self._load_docx(path)
        return []

    def _discover_root_files(self, root: Path) -> list[Path]:
        root = Path(root)
        root_key = str(root)
        with self._cache_lock:
            if not root.exists():
                self._clear_root_cache_locked(root, root_key)
                return []
            snapshot = self._discovery_cache.get(root_key)
            if snapshot is not None and self._snapshot_is_current_locked(snapshot):
                return list(snapshot.files)
            snapshot = self._scan_root_locked(root)
            self._discovery_cache[root_key] = snapshot
            return list(snapshot.files)

    def _snapshot_is_current_locked(self, snapshot: _DiscoverySnapshot) -> bool:
        for directory, cached_mtime_ns in zip(
            snapshot.directories,
            snapshot.directory_mtimes_ns,
            strict=False,
        ):
            try:
                current_mtime_ns = directory.stat().st_mtime_ns
            except OSError:
                return False
            if current_mtime_ns != cached_mtime_ns:
                return False
        return True

    def _scan_root_locked(self, root: Path) -> _DiscoverySnapshot:
        files: list[Path] = []
        directories: list[Path] = []
        directory_mtimes_ns: list[int] = []
        metadata_updates: dict[str, _CachedFileMetadata] = {}

        try:
            root_mtime_ns = root.stat().st_mtime_ns
        except OSError:
            self._clear_root_cache_locked(root, str(root))
            return _DiscoverySnapshot((), (), ())

        stack: list[tuple[Path, int]] = [(root, root_mtime_ns)]
        while stack:
            current_path, current_mtime_ns = stack.pop()
            directories.append(current_path)
            directory_mtimes_ns.append(current_mtime_ns)
            try:
                with os.scandir(current_path) as entries:
                    child_directories: list[tuple[Path, int]] = []
                    for entry in entries:
                        try:
                            if entry.is_dir():
                                child_directories.append(
                                    (Path(entry.path), entry.stat().st_mtime_ns)
                                )
                                continue
                            if not entry.is_file():
                                continue
                        except OSError:
                            continue

                        suffix = Path(entry.name).suffix.lower()
                        if suffix not in SUPPORTED_EXTENSIONS:
                            continue
                        try:
                            stat_result = entry.stat()
                        except OSError:
                            continue

                        path = Path(entry.path)
                        files.append(path)
                        metadata_updates[str(path)] = _CachedFileMetadata(
                            path=path,
                            name=entry.name,
                            suffix=suffix,
                            size=stat_result.st_size,
                            mtime_ns=stat_result.st_mtime_ns,
                        )

                    if child_directories:
                        child_directories.sort(key=lambda item: item[0], reverse=True)
                        stack.extend(child_directories)
            except OSError:
                continue

        files.sort()
        self._replace_root_metadata_locked(root, metadata_updates)
        return _DiscoverySnapshot(
            files=tuple(files),
            directories=tuple(directories),
            directory_mtimes_ns=tuple(directory_mtimes_ns),
        )

    def _replace_root_metadata_locked(
        self,
        root: Path,
        metadata_updates: dict[str, _CachedFileMetadata],
    ) -> None:
        stale_keys = [
            cache_key
            for cache_key, cached in self._file_metadata_cache.items()
            if self._is_relative_to(cached.path, root) and cache_key not in metadata_updates
        ]
        for cache_key in stale_keys:
            self._file_metadata_cache.pop(cache_key, None)
        self._file_metadata_cache.update(metadata_updates)

    def _clear_root_cache_locked(self, root: Path, root_key: str) -> None:
        self._discovery_cache.pop(root_key, None)
        stale_keys = [
            cache_key
            for cache_key, cached in self._file_metadata_cache.items()
            if self._is_relative_to(cached.path, root)
        ]
        for cache_key in stale_keys:
            self._file_metadata_cache.pop(cache_key, None)

    def _load_text(self, path: Path) -> list[DocumentSection]:
        text = path.read_text(encoding="utf-8", errors="replace")
        return [self._build_section(path, text=text, section="Body")]

    def _load_json(self, path: Path) -> list[DocumentSection]:
        payload = json.loads(path.read_text(encoding="utf-8"))
        sections: list[DocumentSection] = []
        append_section = sections.append
        if isinstance(payload, list):
            rows_per_section = self.settings.rag_structured_rows_per_section
            batch_lines: list[str] = []
            append_line = batch_lines.append
            batch_start = 1
            dumps = json.dumps
            for index, item in enumerate(payload, start=1):
                append_line(f"Record {index}: {dumps(item, ensure_ascii=True, default=str)}")
                if len(batch_lines) >= rows_per_section:
                    append_section(
                        self._build_section(
                            path,
                            text="\n".join(batch_lines),
                            section=f"Records {batch_start}-{index}",
                        )
                    )
                    batch_lines.clear()
                    batch_start = index + 1
            if batch_lines:
                append_section(
                    self._build_section(
                        path,
                        text="\n".join(batch_lines),
                        section=f"Records {batch_start}-{batch_start + len(batch_lines) - 1}",
                    )
                )
        else:
            append_section(
                self._build_section(
                    path,
                    text=json.dumps(payload, indent=2, ensure_ascii=True, default=str),
                    section="Document",
                )
            )
        return sections

    def _load_csv(self, path: Path) -> list[DocumentSection]:
        sections: list[DocumentSection] = []
        append_section = sections.append
        rows_per_section = self.settings.rag_structured_rows_per_section
        csv_metadata = {"document_type": "csv"}
        clean_scalar = self._clean_scalar
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.DictReader(handle)
            headers = reader.fieldnames or []
            append_section(
                self._build_section(
                    path,
                    text=f"Dataset columns: {', '.join(headers)}",
                    section="Schema",
                    metadata=csv_metadata,
                )
            )
            batch_lines: list[str] = []
            append_line = batch_lines.append
            batch_start = 1
            for index, row in enumerate(reader, start=1):
                append_line(
                    f"Row {index}: "
                    + ", ".join(
                        f"{key}={clean_scalar(value)}"
                        for key, value in row.items()
                    )
                )
                if len(batch_lines) >= rows_per_section:
                    append_section(
                        self._build_section(
                            path,
                            text="\n".join(batch_lines),
                            section=f"Rows {batch_start}-{index}",
                            metadata=csv_metadata,
                        )
                    )
                    batch_lines.clear()
                    batch_start = index + 1
            if batch_lines:
                append_section(
                    self._build_section(
                        path,
                        text="\n".join(batch_lines),
                        section=f"Rows {batch_start}-{batch_start + len(batch_lines) - 1}",
                        metadata=csv_metadata,
                    )
                )
        return sections

    def _load_xlsx(self, path: Path) -> list[DocumentSection]:
        sections: list[DocumentSection] = []
        append_section = sections.append
        rows_per_section = self.settings.rag_structured_rows_per_section
        clean_scalar = self._clean_scalar
        dumps = json.dumps
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                category=UserWarning,
                module=OPENPYXL_WARNING_MODULE,
            )
            workbook = load_workbook(filename=path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = worksheet.iter_rows(values_only=True)
                header_row = next(rows, None)
                if header_row is None:
                    continue
                headers = [str(value) for value in header_row if value is not None]
                header_count = len(headers)
                sheet_metadata = {"document_type": "xlsx", "sheet": sheet_name}
                append_section(
                    self._build_section(
                        path,
                        text=f"Sheet '{sheet_name}' columns: {', '.join(headers)}",
                        section=f"{sheet_name} schema",
                        page=sheet_name,
                        metadata=sheet_metadata,
                    )
                )
                batch_lines: list[str] = []
                append_line = batch_lines.append
                batch_start = 1
                for index, row in enumerate(rows, start=1):
                    column_limit = min(header_count, len(row))
                    record = {
                        headers[column_index]: clean_scalar(row[column_index])
                        for column_index in range(column_limit)
                    }
                    append_line(f"Row {index}: {dumps(record, ensure_ascii=True, default=str)}")
                    if len(batch_lines) >= rows_per_section:
                        append_section(
                            self._build_section(
                                path,
                                text="\n".join(batch_lines),
                                section=f"{sheet_name} rows {batch_start}-{index}",
                                page=sheet_name,
                                metadata=sheet_metadata,
                            )
                        )
                        batch_lines.clear()
                        batch_start = index + 1
                if batch_lines:
                    append_section(
                        self._build_section(
                            path,
                            text="\n".join(batch_lines),
                            section=f"{sheet_name} rows {batch_start}-{batch_start + len(batch_lines) - 1}",
                            page=sheet_name,
                            metadata=sheet_metadata,
                        )
                    )
        finally:
            workbook.close()
        return sections

    def _load_te_dat(self, path: Path) -> list[DocumentSection]:
        metadata = inspect_te_file(path)
        total_rows = int(metadata["total_rows"])
        preview_rows = min(self.settings.rag_te_preview_rows, total_rows)
        selected_rows = resolve_te_row_target(
            total_rows,
            sample_mode=True,
            sample_fraction=min(1.0, preview_rows / max(total_rows, 1)),
            explicit_limit=preview_rows,
        )
        selected_indices = (
            None if selected_rows >= total_rows else select_te_sample_indices(total_rows, selected_rows)
        )
        sample_lines: list[str] = []
        append_line = sample_lines.append
        feature_pairs = te_feature_names()[:8]
        for chunk in iter_te_row_chunks(
            path,
            selected_indices=selected_indices,
            chunk_size=max(1, min(self.settings.rag_te_preview_rows, 8)),
        ):
            for row_index, row_values in chunk:
                pairs = [
                    f"{feature_name}={round(float(row_values[column_index]), 3)}"
                    for column_index, (feature_name, _) in enumerate(feature_pairs)
                ]
                append_line(f"Observation {row_index + 1}: " + ", ".join(pairs))
                if len(sample_lines) >= preview_rows:
                    break
            if len(sample_lines) >= preview_rows:
                break

        scenario_type = "testing" if path.stem.endswith("_te") else "training"
        text = (
            f"Tennessee Eastman scenario {path.stem} ({scenario_type}). "
            f"Rows={total_rows}. Columns=52. "
            "Feature vector contains 41 measured variables and 11 manipulated variables. "
            "Representative observations: "
            + " ".join(sample_lines)
        )
        return [
            self._build_section(
                path,
                text=text,
                section=f"{path.stem} summary",
                metadata={
                    "document_type": "tennessee_dat",
                    "scenario": path.stem,
                    "scenario_type": scenario_type,
                    "rows": total_rows,
                },
            )
        ]

    def _load_pdf(self, path: Path) -> list[DocumentSection]:
        if PdfReader is None:
            logger.warning("Skipping PDF document without pypdf installed: %s", path)
            return []
        reader = PdfReader(str(path))
        sections: list[DocumentSection] = []
        append_section = sections.append
        for page_number, page in enumerate(reader.pages, start=1):
            text = page.extract_text()
            if not text:
                continue
            stripped = text.strip()
            if not stripped:
                continue
            append_section(
                self._build_section(
                    path,
                    text=text,
                    section=f"Page {page_number}",
                    page=str(page_number),
                    metadata={"document_type": "pdf"},
                )
            )
        return sections

    def _load_docx(self, path: Path) -> list[DocumentSection]:
        if DocxDocument is None:
            logger.warning("Skipping DOCX document without python-docx installed: %s", path)
            return []
        document = DocxDocument(str(path))
        paragraphs: list[str] = []
        append_paragraph = paragraphs.append
        for paragraph in document.paragraphs:
            text = paragraph.text.strip()
            if text:
                append_paragraph(text)
        return [
            self._build_section(
                path,
                text="\n".join(paragraphs),
                section="Document",
                metadata={"document_type": "docx"},
            )
        ]

    def _build_section(
        self,
        path: Path,
        *,
        text: str,
        section: str,
        page: str | None = None,
        metadata: dict[str, Any] | None = None,
    ) -> DocumentSection:
        cached = self._get_file_metadata(path)
        section_metadata = {
            "document_name": cached.name,
            "relative_path": cached.relative_source or self._relative_source(path),
            "extension": cached.suffix,
        }
        if metadata:
            section_metadata.update(metadata)
        return DocumentSection(
            document_name=cached.name,
            source=cached.resolved_source or str(path.resolve()),
            section=section,
            text=text,
            page=page,
            metadata=section_metadata,
        )

    def _get_file_metadata(self, path: Path) -> _CachedFileMetadata:
        key = str(path)
        with self._cache_lock:
            cached = self._file_metadata_cache.get(key)
            if cached is None:
                cached = self._build_file_metadata(path)
                self._file_metadata_cache[key] = cached
            if cached.resolved_source is None or cached.relative_source is None:
                resolved_source = str(path.resolve())
                cached.resolved_source = resolved_source
                cached.relative_source = self._relative_source_from_resolved(resolved_source)
            return cached

    @staticmethod
    def _build_file_metadata(path: Path) -> _CachedFileMetadata:
        try:
            stat_result = path.stat()
            size = stat_result.st_size
            mtime_ns = stat_result.st_mtime_ns
        except OSError:
            size = 0
            mtime_ns = 0
        return _CachedFileMetadata(
            path=path,
            name=path.name,
            suffix=path.suffix.lower(),
            size=size,
            mtime_ns=mtime_ns,
        )

    @staticmethod
    def _clean_scalar(value: Any) -> Any:
        if value is None:
            return None
        if hasattr(value, "isoformat"):
            try:
                return value.isoformat()
            except TypeError:
                return str(value)
        return value

    @staticmethod
    def _relative_source(path: Path) -> str:
        return ModularDocumentLoader._relative_source_from_resolved(str(path.resolve()))

    @staticmethod
    def _relative_source_from_resolved(resolved_source: str) -> str:
        resolved_path = Path(resolved_source)
        try:
            return str(resolved_path.relative_to(BASE_DIR))
        except ValueError:
            return resolved_source

    @staticmethod
    def _is_relative_to(path: Path, root: Path) -> bool:
        try:
            path.relative_to(root)
            return True
        except ValueError:
            return False
